# -*- coding: utf-8 -*-
"""
Created on Fri Nov 30 00:13:19 2018

@author: mdinski
"""
import openpyxl

metin = """
********************************************************
*                                                      *
*           Bibliotheksverfolgungsprogramm             *
*                                                      *
********************************************************
"""

print(metin)

sifre=1071
gr_sifre=input("Bitte Passwort eingeben : ")

if int(gr_sifre) == int(sifre) :
       print("\n*****WILLKOMMEN IN DER bib-BIBLIOTHEK*****\n")
       
else : 
       print(exit)
       exit()
       
              





while True:
    print("(1) Mitglied zur Bibliothek hinzufügen")
    print("(2) Mitglied aus Bibliothek entfernen")
    print("(3) Buch hinzufügen" )
    print("(4) Buch herausnehmen")
    print("(5) Mitglied zum Buchnehmen hinzufügen")
    print("(6) Mitglieder, die Bücher hinterlassen")
    print("(7) Suchen des Buches")
    print("(8) Suchen des Mitgliedes\n")

    islem_giris=input("Geben Sie die gewünschte Transaktionsnummer ein : ")
   
    if int(islem_giris) > 8 :
        print("\n***Du hast falsch eingegeben. Versuch es noch einmal.***\n")
    else : 
           if int(islem_giris) == 1 :
                  üye_isim=input("Vorname des Mitgliedes : ")
                  üye_soyisim=input("Nachname des Mitgliedes :  ")
                  üye_no=input("Manuelle Mitgliednummer : ")
       
                  kitap=openpyxl.load_workbook("üyeler.xlsx")
                  sayfa=kitap.get_sheet_by_name("Üyeler")
                  k=1
                  while True:
                         lst_dgr = sayfa.cell(k,1)
                         if lst_dgr.value == None:
                                sayfa.cell(row=k,column=1,value=üye_no)
                                sayfa.cell(row=k,column=2,value=üye_isim)
                                sayfa.cell(row=k,column=3,value=üye_soyisim)
                                kitap.save("üyeler.xlsx")
                                kitap.close()
                                print ("\nMitglied erfolgreich hinzugefügt!\n")
                                break
                         else:
                                k+=1
                                
                                
                  
                      
           elif int(islem_giris) == 2 : 
                 ck_üyeno=input("Mitgliednummer zum Löschen vom System : ")
                 kitap=openpyxl.load_workbook("üyeler.xlsx")
                 sayfa=kitap.get_sheet_by_name("Üyeler")
                 i=1
                 while i < 9999 :
                        lst_dgr=sayfa.cell(i,1)
                        if lst_dgr.value == ck_üyeno :
                               sayfa.cell(row=i,column=1,value="")
                               sayfa.cell(row=i,column=2,value="")
                               sayfa.cell(row=i,column=3,value="")
                               kitap.save("üyeler.xlsx")
                               kitap.close()
                               print("\nerflogreich gelöscht!\n")
                               break
                       
                        else:
                               i = i+1
                                                     
                        
           elif int(islem_giris) == 3 :
                  ek_kitap=input("Buchname zum Addieren : ")
                  ek_serino=input("Serialnummer vom Buch : ")
                  kitap=openpyxl.load_workbook("kitaplar.xlsx")
                  sayfa=kitap.get_sheet_by_name("Kitaplar")
                  e=1
                  while True : 
                         lst_dgr=sayfa.cell(e,1)
                         if lst_dgr.value == None :
                                sayfa.cell(row=e,column=1,value=ek_kitap)
                                sayfa.cell(row=e,column=2,value=ek_serino)
                                kitap.save("kitaplar.xlsx")
                                kitap.close()
                                print("\nBuch erfolgreich hinzugefügt!\n")
                                break
                         else :
                               e=e+1
                               
                               
                               
           elif int(islem_giris) == 4 :
                  ck_kitap=input("Mitgliednummer zum Löschen vom System : ")
                  kitap=openpyxl.load_workbook("kitaplar.xlsx")
                  sayfa=kitap.get_sheet_by_name("Kitaplar")
                  z=1
                  while z < 9999 :
                         lst_dgr1=sayfa.cell(z,2)
                         if lst_dgr1.value == ck_kitap :
                                sayfa.cell(row=z,column=1,value="")
                                sayfa.cell(row=z,column=2,value="")
                                
                                kitap.save("kitaplar.xlsx")
                                kitap.close()
                                print("\n Buch erfolgreich gelöscht!\n")
                                break
                         else :
                                z=z+1
           elif int(islem_giris) == 5 :
                  ktp_üye=input("Mitgliednummer zum Bucheinnehmen : ")
                  vrln_ktp=input("Serialnummer vom eingegebenen Buch: ")
                  kitap=openpyxl.load_workbook("kitapalan.xlsx")
                  sayfa=kitap.get_sheet_by_name("kitapalan")
                  f=1
                  while True :
                         lst_dgr=sayfa.cell(f,1)
                         if lst_dgr.value == None :
                                sayfa.cell(row=f,column=1,value=ktp_üye)
                                sayfa.cell(row=f,column=2,value=vrln_ktp)
                                kitap.save("kitapalan.xlsx")
                                kitap.close()
                                print("\nBuch erfolgreich engegeben!\n")
                                break
                         else :
                                f=f+1
                               
           elif int(islem_giris) == 6 :
                  ck_ktp=input("Kitap Teslim Eden Üyenin Numarasını Giriniz : ")
                  kitap=openpyxl.load_workbook("kitapalan.xlsx")
                  sayfa=kitap.get_sheet_by_name("kitapalan")
                  a=1
                  while a < 9999 :
                         lst_dgr=sayfa.cell(a,1)
                         if lst_dgr.value == ck_ktp :
                                sayfa.cell(row=a,column=1,value="")
                                sayfa.cell(row=a,column=2,value="")
                                kitap.save("kitapalan.xlsx")
                                kitap.close()
                                print("\nÜyeden Kitap Başarıyla ALINDI!\n")
                                print("Kitabı Teslim Ettiğiniz için TEŞEKKÜRLER!\n")
                                break
                         else :
                                a=a+1
           elif int(islem_giris) == 7 :
                  aranacak_serino=input("Aranacak Kitabın Seri Numarasını Giriniz : ")
                  kitap=openpyxl.load_workbook("kitapalan.xlsx")
                  sayfa=kitap.get_sheet_by_name("kitapalan")
                  u=1
                  g=1
                  while u < 9999 :
                         lst_dgr=sayfa.cell(u,2)
                         u=u+1
                         if lst_dgr.value == aranacak_serino :
                                sahip=sayfa.cell(u-1,1)
                                print("\nBu Kitap,",sahip.value," No'lu Üyededir.\n")
                                print("Üyenin Kim Olduğunu Öğrenmek İçin Programdaki 8'Nolu Fonksiyonu Kullanınız.\n")
                                g=0
                                kitap.save("kitapalan.xlsx")
                                kitap.close()
                                break
                  if g == 1 :
                         print("\nBöyle Bir Kitabı Kimse Almamıştır Veya Yanlış Serino Girdiniz.\n")
                         print("Lütfen Kitabın Kütüphanede Olup Olmadığını Kontrol Ediniz.\n")
                         
                                
               
                                
           elif int(islem_giris) == 8 :
                  aranacak_üyeno=input("Aranacak Üyenin Numarasını Giriniz : ")
                  kitap=openpyxl.load_workbook("üyeler.xlsx")
                  sayfa=kitap.get_sheet_by_name("Üyeler")
                  l=1
                  gr=1
                  while l < 9999 :
                         lst_dgr=sayfa.cell(l,1)
                         l=l+1
                         if lst_dgr.value == aranacak_üyeno :
                                oisim=sayfa.cell(l-1,2)
                                osoyisim=sayfa.cell(l-1,3)
                                print("\nBu Üye No'ya Sahip İsim : ",oisim.value,osoyisim.value," 'dir.\n")
                                gr=0
                                kitap.save("üyeler.xlsx")
                                kitap.close()
                                break
                  if gr == 1 :
                         print("\nGirmiş Olduğunuz Numaraya Ait Üye Bulunamamıştır.\n")
                         print("Doğru Girdiğinizden Lütfen Emin Olunuz.\n")
                        
                   
                 
                   
                                
                               
                               
                 

                 
            
                  
             